"""
Motor de Amostragem - Instituto Amostral MVP
Gera planos amostrais com base em dados do TSE e IBGE.
"""

import os
import math
import json
import pandas as pd
from datetime import datetime
from io import BytesIO

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DADOS_DIR = os.path.join(BASE_DIR, "dados")
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")

os.makedirs(OUTPUTS_DIR, exist_ok=True)

CORES = {
    "primaria": colors.HexColor("#1a3a5c"),
    "secundaria": colors.HexColor("#2e7d9e"),
    "acento": colors.HexColor("#e8a020"),
    "claro": colors.HexColor("#f0f4f8"),
    "branco": colors.white,
    "cinza": colors.HexColor("#6b7280"),
    "cinza_claro": colors.HexColor("#e5e7eb"),
    "verde": colors.HexColor("#16a34a"),
    "vermelho": colors.HexColor("#dc2626"),
}


# ─────────────────────────────────────────────────────────────────────────────
# CÁLCULO ESTATÍSTICO
# ─────────────────────────────────────────────────────────────────────────────

Z_MAP = {0.90: 1.645, 0.95: 1.96, 0.99: 2.576}

# Parâmetros operacionais de mercado (institutos de pesquisa)
DEFF_PADRAO_MERCADO = 1.3
TAXA_RESPOSTA_PADRAO = 0.80
MIN_ENTREVISTAS_MUNICIPAL = 400
MIN_ENTREVISTAS_POR_ZONA = 12
ARREDONDAMENTO_AMOSTRA = 10

ORDEM_INSTRUCAO = [
    "Analfabeto",
    "Lê e escreve",
    "Ensino fundamental",
    "Ensino médio",
    "Superior",
]

ORDEM_FAIXA = [
    "De 16 a 24 anos",
    "25 a 34 anos",
    "35 a 44 anos",
    "45 a 59 anos",
    "Acima de 60 anos",
]


def calcular_amostra_minima(N: int, confianca: float = 0.95, margem_erro: float = 0.05) -> int:
    """
    Calcula tamanho mínimo de amostra para população finita.

    Fórmula de Cochran corrigida para população finita:
        n = (Z² × p × q × N) / (e² × (N-1) + Z² × p × q)

    Args:
        N: Tamanho da população (eleitores cadastrados no TSE)
        confianca: Nível de confiança (0.90, 0.95, 0.99)
        margem_erro: Margem de erro (ex: 0.05 = 5%)

    Returns:
        Tamanho mínimo de amostra (inteiro, arredondado para cima)
    """
    Z = Z_MAP.get(confianca, 1.96)
    p = 0.5   # máxima variância (situação mais conservadora)
    q = 1 - p
    e = margem_erro

    n_infinito = (Z ** 2 * p * q) / (e ** 2)
    n = (n_infinito * N) / (n_infinito + N - 1)
    return math.ceil(n)


def calcular_amostra_recomendada(
    N_eleitores: int,
    N_populacao: int,
    n_zonas: int,
    idh: float | None = None,
    confianca: float = 0.95,
    margem_erro: float = 0.05,
) -> dict:
    """
    Calcula o tamanho de amostra recomendado com base nos dados reais do
    município (TSE + IBGE), retornando múltiplos cenários e justificativa.

    Critérios considerados (padrão de mercado):
    - Fórmula de Cochran para população finita (base teórica)
    - Ajuste de design effect (DEFF) operacional
    - Piso mínimo municipal e por zona eleitoral
    - Alvo de campo considerando taxa de resposta
    - Múltiplos cenários de confiança e margem de erro

    Args:
        N_eleitores: Total de eleitores cadastrados (TSE)
        N_populacao: População total do município (IBGE)
        n_zonas: Número de zonas eleitorais
        idh: Mantido por compatibilidade, não usado no dimensionamento
        confianca: Nível de confiança desejado
        margem_erro: Margem de erro desejada

    Returns:
        dict com:
          - recomendado: tamanho final recomendado
          - minimo_cochran: mínimo pela fórmula pura
          - minimo_por_zona: mínimo para cobrir todas as zonas
          - cenarios: lista de cenários alternativos
          - justificativa: texto explicativo
          - parametros: parâmetros usados no cálculo
    """
    # ── 1. Mínimo pela fórmula de Cochran ─────────────────────────────────
    n_cochran = calcular_amostra_minima(N_eleitores, confianca, margem_erro)

    # ── 2. Ajuste DEFF (padrão mercado) ───────────────────────────────────
    deff_aplicado = DEFF_PADRAO_MERCADO
    n_ajustado_deff = math.ceil(n_cochran * deff_aplicado)

    # ── 3. Mínimo por zona (cobertura de todos os estratos) ───────────────
    # Cada zona deve ter ao menos 12 entrevistas para análise intra-zona
    minimo_por_zona = n_zonas * MIN_ENTREVISTAS_POR_ZONA

    # ── 4. Piso municipal de mercado ─────────────────────────────────────
    n_base = max(n_ajustado_deff, minimo_por_zona, MIN_ENTREVISTAS_MUNICIPAL)

    # ── 5. Valor recomendado final (entrevistas completas) ───────────────
    recomendado = math.ceil(n_base / ARREDONDAMENTO_AMOSTRA) * ARREDONDAMENTO_AMOSTRA

    # Alvo operacional de campo (contatos/abordagens), considerando não resposta
    alvo_campo_sugerido = math.ceil(recomendado / TAXA_RESPOSTA_PADRAO)
    alvo_campo_sugerido = math.ceil(alvo_campo_sugerido / ARREDONDAMENTO_AMOSTRA) * ARREDONDAMENTO_AMOSTRA

    # ── 6. Cenários alternativos ──────────────────────────────────────────
    cenarios = []
    for conf, marg, label in [
        (0.90, 0.07, "Econômico (90% / ±7%)"),
        (0.90, 0.05, "Básico (90% / ±5%)"),
        (0.95, 0.05, "Padrão (95% / ±5%) ★"),
        (0.95, 0.04, "Aprimorado (95% / ±4%)"),
        (0.95, 0.03, "Preciso (95% / ±3%)"),
        (0.99, 0.05, "Rigoroso (99% / ±5%)"),
        (0.99, 0.03, "Máximo (99% / ±3%)"),
    ]:
        n_c = calcular_amostra_minima(N_eleitores, conf, marg)
        n_c_adj = max(math.ceil(n_c * deff_aplicado), minimo_por_zona, MIN_ENTREVISTAS_MUNICIPAL)
        n_c_final = math.ceil(n_c_adj / ARREDONDAMENTO_AMOSTRA) * ARREDONDAMENTO_AMOSTRA
        cenarios.append({
            "label": label,
            "confianca": conf,
            "confianca_pct": int(conf * 100),
            "margem_erro": marg,
            "margem_erro_pct": round(marg * 100, 1),
            "n_cochran": n_c,
            "n_recomendado": n_c_final,
            "selecionado": (conf == confianca and marg == margem_erro),
        })

    # ── 7. Margem de erro real com o n recomendado ────────────────────────
    Z = Z_MAP.get(confianca, 1.96)
    p, q = 0.5, 0.5
    # e = Z * sqrt(p*q/n) * sqrt((N-n)/(N-1))  — com fator de correção
    if recomendado < N_eleitores:
        e_real = Z * math.sqrt(p * q / recomendado) * math.sqrt(
            (N_eleitores - recomendado) / (N_eleitores - 1)
        )
    else:
        e_real = 0.0
    margem_real_pct = round(e_real * 100, 2)

    # ── 8. Justificativa textual ──────────────────────────────────────────
    justificativa = (
        f"Fórmula de Cochran (população finita): n₀ = {n_cochran} entrevistas "
        f"para N={N_eleitores:,} eleitores, confiança {int(confianca*100)}% e margem ±{round(margem_erro*100,1)}%. "
        f"Ajuste operacional de mercado (DEFF={deff_aplicado:.2f}) → {n_ajustado_deff}. "
        f"Piso municipal={MIN_ENTREVISTAS_MUNICIPAL} e cobertura mínima por zona ("
        f"{MIN_ENTREVISTAS_POR_ZONA}×{n_zonas}={minimo_por_zona}). "
        f"Valor final recomendado (entrevistas completas), arredondado: {recomendado}. "
        f"Alvo de campo sugerido com taxa de resposta {int(TAXA_RESPOSTA_PADRAO*100)}%: {alvo_campo_sugerido}. "
        f"Margem de erro real estimada para entrevistas completas: ±{margem_real_pct}%."
    ).replace(",", ".")

    return {
        "recomendado": recomendado,
        "alvo_campo_sugerido": alvo_campo_sugerido,
        "minimo_cochran": n_cochran,
        "minimo_por_zona": minimo_por_zona,
        "margem_real_pct": margem_real_pct,
        "cenarios": cenarios,
        "justificativa": justificativa,
        "parametros": {
            "N_eleitores": N_eleitores,
            "N_populacao": N_populacao,
            "n_zonas": n_zonas,
            "idh": idh,
            "confianca": confianca,
            "confianca_pct": int(confianca * 100),
            "margem_erro": margem_erro,
            "margem_erro_pct": round(margem_erro * 100, 1),
            "deff_aplicado": deff_aplicado,
            "taxa_resposta_padrao": TAXA_RESPOSTA_PADRAO,
            "minimo_municipal": MIN_ENTREVISTAS_MUNICIPAL,
            "minimo_por_zona": MIN_ENTREVISTAS_POR_ZONA,
        },
    }


def calcular_quotas(df_zonas: pd.DataFrame, amostra: int) -> pd.DataFrame:
    """
    Distribui a amostra proporcionalmente entre as zonas eleitorais.
    Usa método de Hamilton (maior resto) para garantir soma exata.
    """
    total = df_zonas["ELEITORES_TOTAL"].sum()
    
    # Quota proporcional
    df_zonas = df_zonas.copy()
    df_zonas["PROPORCAO"] = df_zonas["ELEITORES_TOTAL"] / total
    df_zonas["QUOTA_REAL"] = df_zonas["PROPORCAO"] * amostra
    df_zonas["QUOTA"] = df_zonas["QUOTA_REAL"].apply(math.floor)
    
    # Método Hamilton: distribui os restos
    resto = amostra - df_zonas["QUOTA"].sum()
    df_zonas["RESTO"] = df_zonas["QUOTA_REAL"] - df_zonas["QUOTA"]
    indices_maiores_restos = df_zonas["RESTO"].nlargest(int(resto)).index
    df_zonas.loc[indices_maiores_restos, "QUOTA"] += 1
    
    # Quota por gênero
    df_zonas["QUOTA_FEMININO"] = (df_zonas["QUOTA"] * 
        df_zonas["ELEITORES_FEMININO"] / df_zonas["ELEITORES_TOTAL"]).apply(round)
    df_zonas["QUOTA_MASCULINO"] = df_zonas["QUOTA"] - df_zonas["QUOTA_FEMININO"]
    
    return df_zonas


def _normalizar_texto(s: str) -> str:
    if s is None:
        return ""
    return " ".join(str(s).strip().upper().split())


def _alocar_hamilton(total: int, pcts: list[float]) -> list[int]:
    if total <= 0:
        return [0 for _ in pcts]
    soma_pct = sum(pcts)
    if soma_pct <= 0:
        return [0 for _ in pcts]
    quotas_reais = [(pct / soma_pct) * total for pct in pcts]
    quotas = [math.floor(q) for q in quotas_reais]
    resto = total - sum(quotas)
    if resto > 0:
        restos = sorted(
            [(i, quotas_reais[i] - quotas[i]) for i in range(len(quotas))],
            key=lambda x: x[1],
            reverse=True,
        )
        for i, _ in restos[:resto]:
            quotas[i] += 1
    return quotas


def _calcular_percentuais_municipais(df: pd.DataFrame, coluna_categoria: str, ordem: list[str]) -> list[dict]:
    if df.empty:
        return []
    base = (
        df.groupby(coluna_categoria, as_index=False)["QT_ELEITORES"]
        .sum()
        .sort_values("QT_ELEITORES", ascending=False)
    )
    total = int(base["QT_ELEITORES"].sum())
    if total <= 0:
        return []

    mapa = {str(row[coluna_categoria]): int(row["QT_ELEITORES"]) for _, row in base.iterrows()}
    categorias = ordem if ordem else list(mapa.keys())
    out = []
    for cat in categorias:
        qt = mapa.get(cat, 0)
        out.append({"categoria": cat, "pct": round((qt / total) * 100, 6)})
    return out


def _mapear_faixa_etaria_tse(ds: str) -> str | None:
    t = _normalizar_texto(ds)
    if not t:
        return None
    if "16" in t or "17" in t or "18" in t or "19" in t or "20" in t or "21" in t or "22" in t or "23" in t or "24" in t:
        return "De 16 a 24 anos"
    if "25" in t or "26" in t or "27" in t or "28" in t or "29" in t or "30" in t or "31" in t or "32" in t or "33" in t or "34" in t:
        return "25 a 34 anos"
    if "35" in t or "36" in t or "37" in t or "38" in t or "39" in t or "40" in t or "41" in t or "42" in t or "43" in t or "44" in t:
        return "35 a 44 anos"
    if "45" in t or "46" in t or "47" in t or "48" in t or "49" in t or "50" in t or "51" in t or "52" in t or "53" in t or "54" in t or "55" in t or "56" in t or "57" in t or "58" in t or "59" in t:
        return "45 a 59 anos"
    if "60" in t or "61" in t or "62" in t or "63" in t or "64" in t or "65" in t or "66" in t or "67" in t or "68" in t or "69" in t or "70" in t or "71" in t or "72" in t or "73" in t or "74" in t or "75" in t or "79" in t or "80" in t or "84" in t or "89" in t or "94" in t or "95" in t or "99" in t or "100" in t:
        return "Acima de 60 anos"
    return None


def _mapear_instrucao_tse(ds: str) -> str | None:
    t = _normalizar_texto(ds)
    if not t:
        return None
    if "ANALFAB" in t:
        return "Analfabeto"
    if "LE E ESCREVE" in t or "LÊ E ESCREVE" in t:
        return "Lê e escreve"
    if "FUNDAMENTAL" in t:
        return "Ensino fundamental"
    if "MEDIO" in t or "MÉDIO" in t:
        return "Ensino médio"
    if "SUPERIOR" in t:
        return "Superior"
    return None


def _gerar_tabela_dimensao(titulo: str, categorias_pct: list[dict], amostra: int, fonte: str) -> dict:
    categorias = [c["categoria"] for c in categorias_pct]
    pcts = [float(c["pct"]) for c in categorias_pct]
    abs_vals = _alocar_hamilton(amostra, pcts)

    linhas = []
    for i, cat in enumerate(categorias):
        pct_real = round((abs_vals[i] / amostra) * 100, 2) if amostra else 0.0
        linhas.append(
            {
                "categoria": cat,
                "v_absoluto": int(abs_vals[i]),
                "pct": pct_real,
            }
        )

    return {
        "titulo": titulo,
        "fonte": fonte,
        "linhas": linhas,
        "total": {
            "v_absoluto": int(sum(abs_vals)),
            "pct": 100.0,
        },
    }


def calcular_benchmark_estratos(df_mun: pd.DataFrame, amostra_final: int, uf: str, municipio: str) -> dict:
    """
    Calcula estratificação municipal real para o plano de campo, sem perfis
    sintéticos. Usa apenas dados oficiais observados para o município.
    """
    total_eleitores = int(df_mun["ELEITORES_TOTAL"].sum())
    fem = int(df_mun["ELEITORES_FEMININO"].sum())
    masc = int(df_mun["ELEITORES_MASCULINO"].sum())
    fem_pct = round((fem / total_eleitores) * 100, 2) if total_eleitores else 50.0
    masc_pct = round(100.0 - fem_pct, 2)

    tabela_genero = _gerar_tabela_dimensao(
        "GÊNERO",
        [
            {"categoria": "FEMININO", "pct": fem_pct},
            {"categoria": "MASCULINO", "pct": masc_pct},
        ],
        amostra_final,
        "TSE (eleitorado municipal por zona)",
    )

    # Perfil por instrução/faixa etária com base municipal real (TSE)
    caminho_perfil_tse = os.path.join(DADOS_DIR, "tse_perfil.csv")
    tabelas = [tabela_genero]
    observacoes = []

    if os.path.exists(caminho_perfil_tse):
        try:
            df_perfil = pd.read_csv(caminho_perfil_tse, encoding="utf-8-sig")
            mask = (df_perfil["UF"].str.upper() == uf.upper()) & (
                df_perfil["MUNICIPIO"].str.lower() == municipio.lower()
            )
            df_pm = df_perfil[mask].copy()
            if not df_pm.empty:
                if "DIMENSAO" in df_pm.columns and "CATEGORIA" in df_pm.columns:
                    df_i = df_pm[df_pm["DIMENSAO"] == "INSTRUCAO"].copy()
                    df_i["CATEGORIA"] = df_i["CATEGORIA"].apply(_mapear_instrucao_tse)
                    df_i = df_i[df_i["CATEGORIA"].notna()].copy()
                    instr_pct = _calcular_percentuais_municipais(df_i, "CATEGORIA", ORDEM_INSTRUCAO)
                    if instr_pct:
                        tabelas.append(
                            _gerar_tabela_dimensao(
                                "GRAU DE INSTRUÇÃO",
                                instr_pct,
                                amostra_final,
                                "TSE (perfil municipal por seção eleitoral)",
                            )
                        )

                if "DIMENSAO" in df_pm.columns and "CATEGORIA" in df_pm.columns:
                    df_f = df_pm[df_pm["DIMENSAO"] == "FAIXA_ETARIA"].copy()
                    df_f["CATEGORIA"] = df_f["CATEGORIA"].apply(_mapear_faixa_etaria_tse)
                    df_f = df_f[df_f["CATEGORIA"].notna()].copy()
                    faixa_pct = _calcular_percentuais_municipais(df_f, "CATEGORIA", ORDEM_FAIXA)
                    if faixa_pct:
                        tabelas.append(
                            _gerar_tabela_dimensao(
                                "FAIXA ETÁRIA",
                                faixa_pct,
                                amostra_final,
                                "TSE (perfil municipal por seção eleitoral)",
                            )
                        )
            else:
                observacoes.append("Perfil municipal detalhado (instrução/faixa etária) não encontrado para este município.")
        except Exception:
            observacoes.append("Falha ao ler perfil municipal detalhado (tse_perfil.csv).")
    else:
        observacoes.append("Arquivo tse_perfil.csv não encontrado. Rode gerar_dados.py para habilitar estratificação municipal completa.")

    return {
        "metodologia": (
            "Estratificação municipal real: quotas calculadas por método de Hamilton com soma exata da amostra final. "
            "As proporções são derivadas exclusivamente de dados oficiais observados do município (TSE), "
            "sem uso de percentuais fixos de benchmark para dimensionamento do plano."
        ),
        "observacoes": observacoes,
        "tabelas": tabelas,
    }


def gerar_texto_metodologia_institucional(meta: dict, benchmark: dict | None = None) -> str:
    """Retorna texto institucional padronizado para proposta/relatório técnico."""
    confianca = meta.get("confianca_pct", 95)
    margem = meta.get("margem_erro_pct", 5)
    amostra = meta.get("amostra_final", 0)
    total = meta.get("total_eleitores", 0)
    n_zonas = meta.get("n_zonas", 0)
    margem_real = meta.get("margem_real_pct", margem)

    bloco_benchmark = (
        "Complementarmente, o plano incorpora benchmark estratificado de execução de campo "
        "(gênero, grau de instrução e faixa etária), com alocação por maior resto (Hamilton) para "
        "preservar a soma exata das quotas. "
        "Gênero é ancorado no eleitorado municipal observado no TSE; instrução e faixa etária são "
        "ancoradas no perfil detalhado por seção do próprio município, sem uso de percentuais sintéticos."
        if benchmark and benchmark.get("tabelas")
        else ""
    )

    return (
        "Este plano amostral foi estruturado para assegurar robustez estatística, viabilidade operacional "
        "e rastreabilidade metodológica. O dimensionamento da amostra considera população finita e aplica "
        "a fórmula de Cochran com parâmetros conservadores de máxima variância (p=q=0,5), resultando em "
        f"amostra final de {amostra:,} entrevistas para universo de {total:,} eleitores. "
        f"O desenho atende ao nível de confiança de {confianca}% e margem máxima planejada de ±{margem}%, "
        f"com margem efetiva estimada em ±{margem_real}%. "
        "A seleção é estratificada por zona eleitoral, com distribuição proporcional e ajuste por maior resto, "
        f"garantindo cobertura integral dos {n_zonas} estratos e aderência ao total amostral. "
        f"{bloco_benchmark} "
        "As fontes utilizadas são oficiais e públicas, com referência temporal explicitada em metadados, "
        "permitindo auditoria técnica e reprodutibilidade da entrega."
    ).replace(",", ".")


# ─────────────────────────────────────────────────────────────────────────────
# FUNÇÃO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def gerar_plano(
    uf: str,
    municipio: str,
    amostra: int = None,   # None = calcular automaticamente
    formato: str = "pdf",
    confianca: float = 0.95,
    margem_erro: float = 0.05
) -> dict:
    """
    Gera plano amostral completo para um município.
    
    Returns:
        dict com caminhos dos arquivos gerados e metadados
    """
    # ── Carrega dados ──────────────────────────────────────────────────────
    caminho_tse = os.path.join(DADOS_DIR, "tse.csv")
    caminho_ibge = os.path.join(DADOS_DIR, "ibge.csv")
    
    if not os.path.exists(caminho_tse) or not os.path.exists(caminho_ibge):
        raise FileNotFoundError(
            "Dados não encontrados. Execute 'python gerar_dados.py' primeiro."
        )
    
    df_tse = pd.read_csv(caminho_tse, encoding="utf-8-sig")
    df_ibge = pd.read_csv(caminho_ibge, encoding="utf-8-sig")
    
    # ── Filtra município ───────────────────────────────────────────────────
    uf_upper = uf.upper().strip()
    municipio_strip = municipio.strip()
    
    mask = (df_tse["UF"] == uf_upper) & (
        df_tse["MUNICIPIO"].str.lower() == municipio_strip.lower()
    )
    df_mun = df_tse[mask].copy()
    
    if df_mun.empty:
        raise ValueError(
            f"Município '{municipio}' não encontrado no estado '{uf}'. "
            "Verifique o nome exato."
        )
    
    # ── Dados IBGE ─────────────────────────────────────────────────────────
    mask_ibge = (df_ibge["UF"] == uf_upper) & (
        df_ibge["MUNICIPIO"].str.lower() == municipio_strip.lower()
    )
    df_ibge_mun = df_ibge[mask_ibge]
    ibge_data = df_ibge_mun.iloc[0].to_dict() if not df_ibge_mun.empty else {}
    
    # ── Cálculos ───────────────────────────────────────────────────────────
    total_eleitores = int(df_mun["ELEITORES_TOTAL"].sum())
    n_zonas = len(df_mun)
    pop_total = int(ibge_data.get("POPULACAO_TOTAL", total_eleitores))
    idh_raw = ibge_data.get("IDH")
    try:
        idh = float(idh_raw) if idh_raw is not None and str(idh_raw).strip() != "" else None
    except (TypeError, ValueError):
        idh = None

    # Calcula amostra recomendada com todos os critérios
    calc = calcular_amostra_recomendada(
        N_eleitores=total_eleitores,
        N_populacao=pop_total,
        n_zonas=n_zonas,
        idh=idh,
        confianca=confianca,
        margem_erro=margem_erro,
    )

    amostra_recomendada = calc["recomendado"]
    amostra_minima = calc["minimo_cochran"]

    # Se o usuário não informou amostra, usa a recomendada
    # Se informou, respeita desde que seja >= mínimo de Cochran
    if amostra is None:
        amostra_final = amostra_recomendada
        amostra_solicitada = None
    else:
        amostra_solicitada = amostra
        amostra_final = max(int(amostra), amostra_minima)

    df_quotas = calcular_quotas(df_mun, amostra_final)
    benchmark = calcular_benchmark_estratos(df_mun, amostra_final, uf_upper, municipio_strip)
    metodologia_institucional = gerar_texto_metodologia_institucional(
        {
            "confianca_pct": int(confianca * 100),
            "margem_erro_pct": round(margem_erro * 100, 1),
            "amostra_final": amostra_final,
            "total_eleitores": total_eleitores,
            "n_zonas": len(df_quotas),
            "margem_real_pct": calc["margem_real_pct"],
        },
        benchmark=benchmark,
    )

    # ── Metadados do plano ─────────────────────────────────────────────────
    meta = {
        "uf": uf_upper,
        "municipio": municipio_strip,
        "data_geracao": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "total_eleitores": total_eleitores,
        "amostra_solicitada": amostra_solicitada,
        "amostra_recomendada": amostra_recomendada,
        "amostra_minima": amostra_minima,
        "amostra_final": amostra_final,
        "modo_calculo": "automatico" if amostra_solicitada is None else "manual",
        "confianca_pct": int(confianca * 100),
        "margem_erro_pct": round(margem_erro * 100, 1),
        "margem_real_pct": calc["margem_real_pct"],
        "n_zonas": len(df_quotas),
        "calculo_detalhado": calc,
        "ibge": ibge_data,
        "benchmark": benchmark,
        "metodologia_institucional": metodologia_institucional,
    }
    
    # ── Gera arquivos ──────────────────────────────────────────────────────
    slug = f"{uf_upper}_{municipio_strip.replace(' ', '_')}"
    resultado = {"meta": meta}
    
    # Sempre gera Excel
    caminho_excel = os.path.join(OUTPUTS_DIR, f"{slug}_plano.xlsx")
    gerar_excel(df_quotas, meta, caminho_excel, benchmark=benchmark)
    resultado["excel"] = caminho_excel

    # Detalhamento por zona para o frontend
    resultado["zonas"] = df_quotas[
        [
            "ZONA",
            "ELEITORES_TOTAL",
            "ELEITORES_FEMININO",
            "ELEITORES_MASCULINO",
            "SECOES",
            "QUOTA",
            "QUOTA_FEMININO",
            "QUOTA_MASCULINO",
        ]
    ].to_dict("records")
    resultado["benchmark"] = benchmark
    
    if formato == "pdf":
        caminho_pdf = os.path.join(OUTPUTS_DIR, f"{slug}_plano.pdf")
        gerar_pdf(df_quotas, meta, caminho_pdf, benchmark=benchmark)
        resultado["pdf"] = caminho_pdf
    
    if formato == "markdown" or formato == "md":
        caminho_md = os.path.join(OUTPUTS_DIR, f"{slug}_plano.md")
        gerar_markdown(df_quotas, meta, caminho_md, benchmark=benchmark)
        resultado["markdown"] = caminho_md
    
    return resultado


# ─────────────────────────────────────────────────────────────────────────────
# GERAÇÃO DE EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def gerar_excel(df: pd.DataFrame, meta: dict, caminho: str, benchmark: dict | None = None):
    """Gera planilha Excel formatada com o plano amostral."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # ── Aba 1: Plano Amostral ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Plano Amostral"
    
    # Cores
    cor_header = "1A3A5C"
    cor_sub = "2E7D9E"
    cor_acento = "E8A020"
    cor_claro = "F0F4F8"
    cor_alt = "DBEAFE"
    
    def estilo_celula(ws, linha, col, valor, negrito=False, cor_fundo=None, 
                       cor_fonte="000000", alinhamento="left", tamanho=10):
        cell = ws.cell(row=linha, column=col, value=valor)
        cell.font = Font(bold=negrito, color=cor_fonte, size=tamanho, name="Calibri")
        if cor_fundo:
            cell.fill = PatternFill("solid", fgColor=cor_fundo)
        cell.alignment = Alignment(horizontal=alinhamento, vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return cell
    
    # Título
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"PLANO AMOSTRAL — {meta['municipio'].upper()} / {meta['uf']}"
    cell.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    cell.fill = PatternFill("solid", fgColor=cor_header)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Subtítulo
    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = f"Instituto Amostral  |  Gerado em: {meta['data_geracao']}"
    cell.font = Font(italic=True, size=10, color="FFFFFF", name="Calibri")
    cell.fill = PatternFill("solid", fgColor=cor_sub)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # Resumo estatístico
    ws.row_dimensions[3].height = 10
    
    resumo_dados = [
        ("Total de Eleitores", f"{meta['total_eleitores']:,}".replace(",", ".")),
        ("Amostra Mínima Calculada", f"{meta['amostra_minima']:,}".replace(",", ".")),
        ("Amostra Final Aplicada", f"{meta['amostra_final']:,}".replace(",", ".")),
        ("Nível de Confiança", f"{meta['confianca_pct']}%"),
        ("Margem de Erro", f"±{meta['margem_erro_pct']}%"),
        ("Número de Zonas", str(meta['n_zonas'])),
    ]
    
    ws.merge_cells("A4:H4")
    cell = ws["A4"]
    cell.value = "RESUMO ESTATÍSTICO"
    cell.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    cell.fill = PatternFill("solid", fgColor=cor_acento)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 22
    
    for i, (label, valor) in enumerate(resumo_dados):
        linha = 5 + i
        estilo_celula(ws, linha, 1, label, negrito=True, cor_fundo=cor_claro, tamanho=10)
        ws.merge_cells(f"A{linha}:D{linha}")
        estilo_celula(ws, linha, 5, valor, cor_fundo="FFFFFF", alinhamento="center", tamanho=10)
        ws.merge_cells(f"E{linha}:H{linha}")
        ws.row_dimensions[linha].height = 18
    
    # Espaço
    linha_inicio_tabela = 5 + len(resumo_dados) + 2
    
    # Cabeçalho da tabela
    colunas = [
        "Zona", "Eleitores\nTotal", "Eleitores\nFeminino", "Eleitores\nMasculino",
        "Seções", "Quota\nTotal", "Quota\nFeminino", "Quota\nMasculino"
    ]
    
    ws.merge_cells(f"A{linha_inicio_tabela - 1}:H{linha_inicio_tabela - 1}")
    cell = ws.cell(row=linha_inicio_tabela - 1, column=1, value="DISTRIBUIÇÃO POR ZONA ELEITORAL")
    cell.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    cell.fill = PatternFill("solid", fgColor=cor_header)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[linha_inicio_tabela - 1].height = 22
    
    for j, col_nome in enumerate(colunas):
        cell = estilo_celula(ws, linha_inicio_tabela, j + 1, col_nome, 
                              negrito=True, cor_fundo=cor_sub, cor_fonte="FFFFFF",
                              alinhamento="center", tamanho=9)
        ws.row_dimensions[linha_inicio_tabela].height = 30
    
    # Dados
    cols_df = ["ZONA", "ELEITORES_TOTAL", "ELEITORES_FEMININO", "ELEITORES_MASCULINO",
               "SECOES", "QUOTA", "QUOTA_FEMININO", "QUOTA_MASCULINO"]
    
    for i, (_, row) in enumerate(df.iterrows()):
        linha = linha_inicio_tabela + 1 + i
        cor_bg = cor_alt if i % 2 == 0 else "FFFFFF"
        for j, col in enumerate(cols_df):
            val = row.get(col, "")
            if isinstance(val, float):
                val = int(val)
            estilo_celula(ws, linha, j + 1, val, cor_fundo=cor_bg, alinhamento="center")
        ws.row_dimensions[linha].height = 16
    
    # Totais
    linha_total = linha_inicio_tabela + 1 + len(df)
    totais = ["TOTAL", df["ELEITORES_TOTAL"].sum(), df["ELEITORES_FEMININO"].sum(),
              df["ELEITORES_MASCULINO"].sum(), df["SECOES"].sum(),
              df["QUOTA"].sum(), df["QUOTA_FEMININO"].sum(), df["QUOTA_MASCULINO"].sum()]
    for j, val in enumerate(totais):
        if isinstance(val, float):
            val = int(val)
        estilo_celula(ws, linha_total, j + 1, val, negrito=True, 
                       cor_fundo=cor_acento, cor_fonte="FFFFFF", alinhamento="center")
    ws.row_dimensions[linha_total].height = 20
    
    # Larguras das colunas
    larguras = [12, 14, 14, 14, 10, 12, 14, 14]
    for i, larg in enumerate(larguras):
        ws.column_dimensions[get_column_letter(i + 1)].width = larg
    
    # ── Aba 2: Dados IBGE ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Dados IBGE")
    
    ws2.merge_cells("A1:D1")
    cell = ws2["A1"]
    cell.value = f"DADOS SOCIOECONÔMICOS — {meta['municipio'].upper()} / {meta['uf']}"
    cell.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    cell.fill = PatternFill("solid", fgColor=cor_header)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30
    
    ibge = meta.get("ibge", {})

    def fmt_int_opt(valor):
        if valor is None or str(valor).strip() == "":
            return "N/D"
        try:
            return f"{int(float(valor)):,}".replace(",", ".")
        except (TypeError, ValueError):
            return "N/D"

    def fmt_float_opt(valor, casas=1, prefixo=""):
        if valor is None or str(valor).strip() == "":
            return "N/D"
        try:
            return f"{prefixo}{float(valor):,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except (TypeError, ValueError):
            return "N/D"

    ibge_campos = [
        ("População Total", fmt_int_opt(ibge.get("POPULACAO_TOTAL"))),
        ("ID IBGE", fmt_int_opt(ibge.get("ID_IBGE"))),
        ("IDH Municipal", ibge.get("IDH") if ibge.get("IDH") not in [None, ""] else "N/D"),
        ("PIB per Capita (R$)", fmt_float_opt(ibge.get("PIB_PER_CAPITA"), casas=2, prefixo="R$ ")),
        ("Densidade (hab/km²)", fmt_float_opt(ibge.get("DENSIDADE_HAB_KM2"), casas=1)),
    ]
    
    for i, (label, valor) in enumerate(ibge_campos):
        linha = 3 + i
        estilo_celula(ws2, linha, 1, label, negrito=True, cor_fundo=cor_claro)
        ws2.merge_cells(f"A{linha}:B{linha}")
        estilo_celula(ws2, linha, 3, valor, alinhamento="center")
        ws2.merge_cells(f"C{linha}:D{linha}")
        ws2.row_dimensions[linha].height = 18
    
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 15
    
    # ── Aba 3: Benchmark (entrega estratificada) ──────────────────────────
    if benchmark:
        ws3 = wb.create_sheet("Benchmark")
        ws3.merge_cells("A1:C1")
        t = ws3["A1"]
        t.value = f"BENCHMARK ESTRATIFICADO — {meta['municipio'].upper()} / {meta['uf']}"
        t.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
        t.fill = PatternFill("solid", fgColor=cor_header)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 28

        ws3.merge_cells("A2:C2")
        s = ws3["A2"]
        s.value = benchmark.get("metodologia", "")
        s.font = Font(size=9, color="333333", name="Calibri")
        s.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws3.row_dimensions[2].height = 32

        linha = 4
        cab = ["Categoria", "V. Absoluto", "%"]
        for tabela in benchmark.get("tabelas", []):
            ws3.merge_cells(f"A{linha}:C{linha}")
            ch = ws3[f"A{linha}"]
            ch.value = f"{tabela.get('titulo', '')}  |  Fonte: {tabela.get('fonte', '')}"
            ch.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
            ch.fill = PatternFill("solid", fgColor=cor_sub)
            ch.alignment = Alignment(horizontal="left", vertical="center")
            ws3.row_dimensions[linha].height = 20
            linha += 1

            for j, col_nome in enumerate(cab):
                estilo_celula(ws3, linha, j + 1, col_nome, negrito=True, cor_fundo=cor_acento, cor_fonte="FFFFFF", alinhamento="center", tamanho=9)
            ws3.row_dimensions[linha].height = 18
            linha += 1

            for item in tabela.get("linhas", []):
                vals = [
                    item.get("categoria", ""),
                    int(item.get("v_absoluto", 0)),
                    f"{float(item.get('pct', 0)):.2f}%",
                ]
                for j, v in enumerate(vals):
                    estilo_celula(ws3, linha, j + 1, v, cor_fundo="FFFFFF" if linha % 2 else cor_claro, alinhamento="center" if j else "left", tamanho=9)
                ws3.row_dimensions[linha].height = 17
                linha += 1

            total = tabela.get("total", {})
            vals_t = [
                "TOTAL",
                int(total.get("v_absoluto", 0)),
                "100,00%",
            ]
            for j, v in enumerate(vals_t):
                estilo_celula(ws3, linha, j + 1, v, negrito=True, cor_fundo=cor_alt, alinhamento="center" if j else "left", tamanho=9)
            ws3.row_dimensions[linha].height = 18
            linha += 2

        if benchmark.get("observacoes"):
            ws3.merge_cells(f"A{linha}:C{linha}")
            o = ws3[f"A{linha}"]
            o.value = "Observações: " + " | ".join(benchmark.get("observacoes", []))
            o.font = Font(size=9, color="666666", name="Calibri")
            o.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws3.row_dimensions[linha].height = 34

        larguras3 = [36, 14, 12]
        for i, larg in enumerate(larguras3):
            ws3.column_dimensions[get_column_letter(i + 1)].width = larg

    wb.save(caminho)


# ─────────────────────────────────────────────────────────────────────────────
# GERAÇÃO DE PDF
# ─────────────────────────────────────────────────────────────────────────────

def gerar_pdf(df: pd.DataFrame, meta: dict, caminho: str, benchmark: dict | None = None):
    """Gera relatório PDF profissional com o plano amostral."""
    
    doc = SimpleDocTemplate(
        caminho,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2.5*cm,
        bottomMargin=2*cm,
    )
    
    estilos = getSampleStyleSheet()
    
    # Estilos customizados
    estilo_titulo = ParagraphStyle(
        "Titulo",
        parent=estilos["Title"],
        fontSize=18,
        textColor=CORES["primaria"],
        spaceAfter=6,
        fontName="Helvetica-Bold",
        alignment=TA_CENTER,
    )
    estilo_subtitulo = ParagraphStyle(
        "Subtitulo",
        parent=estilos["Normal"],
        fontSize=11,
        textColor=CORES["secundaria"],
        spaceAfter=4,
        fontName="Helvetica",
        alignment=TA_CENTER,
    )
    estilo_secao = ParagraphStyle(
        "Secao",
        parent=estilos["Heading2"],
        fontSize=12,
        textColor=CORES["branco"],
        backColor=CORES["primaria"],
        spaceBefore=12,
        spaceAfter=8,
        fontName="Helvetica-Bold",
        leftIndent=-10,
        rightIndent=-10,
        borderPadding=(4, 8, 4, 8),
    )
    estilo_normal = ParagraphStyle(
        "Normal2",
        parent=estilos["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#374151"),
        fontName="Helvetica",
    )
    estilo_rodape = ParagraphStyle(
        "Rodape",
        parent=estilos["Normal"],
        fontSize=8,
        textColor=CORES["cinza"],
        alignment=TA_CENTER,
        fontName="Helvetica-Oblique",
    )
    
    elementos = []
    
    # ── Cabeçalho ──────────────────────────────────────────────────────────
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph("INSTITUTO AMOSTRAL", estilo_subtitulo))
    elementos.append(Paragraph(
        f"PLANO AMOSTRAL — {meta['municipio'].upper()} / {meta['uf']}",
        estilo_titulo
    ))
    elementos.append(Paragraph(
        f"Gerado em: {meta['data_geracao']}  |  Nível de Confiança: {meta['confianca_pct']}%  |  Margem de Erro: ±{meta['margem_erro_pct']}%",
        estilo_subtitulo
    ))
    elementos.append(HRFlowable(width="100%", thickness=2, color=CORES["acento"], spaceAfter=12))
    
    # ── Resumo Estatístico ─────────────────────────────────────────────────
    elementos.append(Paragraph("RESUMO ESTATÍSTICO", estilo_secao))
    elementos.append(Spacer(1, 0.3*cm))
    
    dados_resumo = [
        ["Indicador", "Valor"],
        ["Total de Eleitores Cadastrados", f"{meta['total_eleitores']:,}".replace(",", ".")],
        ["Amostra Mínima Calculada (Cochran)", f"{meta['amostra_minima']:,}".replace(",", ".")],
        ["Amostra Final Aplicada", f"{meta['amostra_final']:,}".replace(",", ".")],
        ["Número de Zonas Eleitorais", str(meta['n_zonas'])],
        ["Nível de Confiança", f"{meta['confianca_pct']}%"],
        ["Margem de Erro Máxima", f"±{meta['margem_erro_pct']}%"],
    ]
    
    tabela_resumo = Table(dados_resumo, colWidths=[10*cm, 6*cm])
    tabela_resumo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), CORES["primaria"]),
        ("TEXTCOLOR", (0, 0), (-1, 0), CORES["branco"]),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [CORES["claro"], CORES["branco"]]),
        ("GRID", (0, 0), (-1, -1), 0.5, CORES["cinza_claro"]),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("FONTNAME", (0, 2), (0, 2), "Helvetica-Bold"),
        ("TEXTCOLOR", (1, 2), (1, 2), CORES["secundaria"]),
        ("FONTNAME", (0, 3), (0, 3), "Helvetica-Bold"),
        ("TEXTCOLOR", (1, 3), (1, 3), CORES["verde"]),
    ]))
    elementos.append(tabela_resumo)
    elementos.append(Spacer(1, 0.5*cm))
    
    # ── Dados IBGE ─────────────────────────────────────────────────────────
    ibge = meta.get("ibge", {})
    if ibge:
        def fmt_int_opt(valor):
            if valor is None or str(valor).strip() == "":
                return "N/D"
            try:
                return f"{int(float(valor)):,}".replace(",", ".")
            except (TypeError, ValueError):
                return "N/D"

        def fmt_float_opt(valor, casas=1, prefixo=""):
            if valor is None or str(valor).strip() == "":
                return "N/D"
            try:
                return f"{prefixo}{float(valor):,.{casas}f}".replace(",", ".")
            except (TypeError, ValueError):
                return "N/D"

        elementos.append(Paragraph("PERFIL SOCIOECONÔMICO (IBGE)", estilo_secao))
        elementos.append(Spacer(1, 0.3*cm))

        dados_ibge = [
            ["Indicador", "Valor", "Indicador", "Valor"],
            ["População Total", fmt_int_opt(ibge.get("POPULACAO_TOTAL")),
             "ID IBGE", fmt_int_opt(ibge.get("ID_IBGE"))],
            ["IDH Municipal", ibge.get("IDH") if ibge.get("IDH") not in [None, ""] else "N/D",
             "PIB per Capita", fmt_float_opt(ibge.get("PIB_PER_CAPITA"), casas=0, prefixo="R$ ")],
            ["Densidade (hab/km²)", fmt_float_opt(ibge.get("DENSIDADE_HAB_KM2"), casas=1),
             "Fonte", "IBGE (API pública)"],
        ]
        
        tabela_ibge = Table(dados_ibge, colWidths=[5*cm, 4*cm, 5*cm, 4*cm])
        tabela_ibge.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), CORES["secundaria"]),
            ("TEXTCOLOR", (0, 0), (-1, 0), CORES["branco"]),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [CORES["claro"], CORES["branco"]]),
            ("GRID", (0, 0), (-1, -1), 0.5, CORES["cinza_claro"]),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 1), (2, -1), "Helvetica-Bold"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elementos.append(tabela_ibge)
        elementos.append(Spacer(1, 0.5*cm))
    
    # ── Tabela de Quotas por Zona ──────────────────────────────────────────
    elementos.append(Paragraph("DISTRIBUIÇÃO AMOSTRAL POR ZONA ELEITORAL", estilo_secao))
    elementos.append(Spacer(1, 0.3*cm))
    
    colunas_tabela = [
        "Zona", "Eleitores\nTotal", "Eleit.\nFem.", "Eleit.\nMasc.",
        "Seções", "Quota\nTotal", "Quota\nFem.", "Quota\nMasc.", "Proporção"
    ]
    
    dados_tabela = [colunas_tabela]
    total_eleitores = df["ELEITORES_TOTAL"].sum()
    
    for _, row in df.iterrows():
        proporcao = f"{(row['ELEITORES_TOTAL'] / total_eleitores * 100):.1f}%"
        dados_tabela.append([
            str(row["ZONA"]),
            f"{int(row['ELEITORES_TOTAL']):,}".replace(",", "."),
            f"{int(row['ELEITORES_FEMININO']):,}".replace(",", "."),
            f"{int(row['ELEITORES_MASCULINO']):,}".replace(",", "."),
            str(int(row["SECOES"])),
            str(int(row["QUOTA"])),
            str(int(row["QUOTA_FEMININO"])),
            str(int(row["QUOTA_MASCULINO"])),
            proporcao,
        ])
    
    # Linha de totais
    dados_tabela.append([
        "TOTAL",
        f"{int(df['ELEITORES_TOTAL'].sum()):,}".replace(",", "."),
        f"{int(df['ELEITORES_FEMININO'].sum()):,}".replace(",", "."),
        f"{int(df['ELEITORES_MASCULINO'].sum()):,}".replace(",", "."),
        str(int(df["SECOES"].sum())),
        str(int(df["QUOTA"].sum())),
        str(int(df["QUOTA_FEMININO"].sum())),
        str(int(df["QUOTA_MASCULINO"].sum())),
        "100%",
    ])
    
    col_widths = [1.5*cm, 2.5*cm, 2*cm, 2*cm, 1.5*cm, 2*cm, 2*cm, 2*cm, 2*cm]
    tabela_zonas = Table(dados_tabela, colWidths=col_widths, repeatRows=1)
    
    n_linhas = len(dados_tabela)
    tabela_zonas.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), CORES["primaria"]),
        ("TEXTCOLOR", (0, 0), (-1, 0), CORES["branco"]),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7.5),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -2), 7.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [CORES["claro"], CORES["branco"]]),
        ("BACKGROUND", (0, -1), (-1, -1), CORES["acento"]),
        ("TEXTCOLOR", (0, -1), (-1, -1), CORES["branco"]),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, CORES["cinza_claro"]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    elementos.append(tabela_zonas)

    # ── Benchmark estratificado ────────────────────────────────────────────
    if benchmark and benchmark.get("tabelas"):
        elementos.append(PageBreak())
        elementos.append(Paragraph("BENCHMARK ESTRATIFICADO DE ENTREGA", estilo_secao))
        elementos.append(Spacer(1, 0.2*cm))
        elementos.append(Paragraph(benchmark.get("metodologia", ""), estilo_normal))
        elementos.append(Spacer(1, 0.3*cm))

        for tabela in benchmark.get("tabelas", []):
            elementos.append(Paragraph(f"<b>{tabela.get('titulo','')}</b> — {tabela.get('fonte','')}", estilo_normal))
            dados_b = [["Categoria", "V.Abs", "%"]]
            for item in tabela.get("linhas", []):
                dados_b.append([
                    item.get("categoria", ""),
                    str(int(item.get("v_absoluto", 0))),
                    f"{float(item.get('pct', 0)):.2f}%",
                ])
            tot = tabela.get("total", {})
            dados_b.append([
                "TOTAL",
                str(int(tot.get("v_absoluto", 0))),
                "100,00%",
            ])

            tb = Table(dados_b, colWidths=[9.2*cm, 2.5*cm, 2.3*cm], repeatRows=1)
            tb.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), CORES["secundaria"]),
                ("TEXTCOLOR", (0, 0), (-1, 0), CORES["branco"]),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.3, CORES["cinza_claro"]),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [CORES["claro"], CORES["branco"]]),
                ("BACKGROUND", (0, -1), (-1, -1), CORES["acento"]),
                ("TEXTCOLOR", (0, -1), (-1, -1), CORES["branco"]),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.6),
            ]))
            elementos.append(tb)
            elementos.append(Spacer(1, 0.25*cm))

        if benchmark.get("observacoes"):
            elementos.append(Paragraph("Observações: " + " | ".join(benchmark.get("observacoes", [])), estilo_rodape))
            elementos.append(Spacer(1, 0.2*cm))
    
    # ── Metodologia ────────────────────────────────────────────────────────
    elementos.append(Spacer(1, 0.8*cm))
    elementos.append(Paragraph("NOTA METODOLÓGICA", estilo_secao))
    elementos.append(Spacer(1, 0.3*cm))
    
    nota = meta.get("metodologia_institucional") or gerar_texto_metodologia_institucional(meta, benchmark=benchmark)
    elementos.append(Paragraph(nota, estilo_normal))
    
    # ── Rodapé ─────────────────────────────────────────────────────────────
    elementos.append(Spacer(1, 1*cm))
    elementos.append(HRFlowable(width="100%", thickness=1, color=CORES["cinza_claro"]))
    elementos.append(Spacer(1, 0.2*cm))
    elementos.append(Paragraph(
        f"Instituto Amostral  •  Plano Amostral MVP  •  {meta['data_geracao']}  •  Documento gerado automaticamente",
        estilo_rodape
    ))
    
    doc.build(elementos)


# ─────────────────────────────────────────────────────────────────────────────
# GERAÇÃO DE MARKDOWN
# ─────────────────────────────────────────────────────────────────────────────

def gerar_markdown(df: pd.DataFrame, meta: dict, caminho: str, benchmark: dict | None = None):
    """Gera relatório em Markdown com o plano amostral."""
    
    linhas = []
    linhas.append(f"# Plano Amostral — {meta['municipio']} / {meta['uf']}")
    linhas.append(f"\n> **Instituto Amostral** | Gerado em: {meta['data_geracao']}\n")
    linhas.append("---\n")
    
    linhas.append("## Resumo Estatístico\n")
    linhas.append("| Indicador | Valor |")
    linhas.append("|-----------|-------|")
    linhas.append(f"| Total de Eleitores | {meta['total_eleitores']:,} |".replace(",", "."))
    linhas.append(f"| Amostra Mínima (Cochran) | {meta['amostra_minima']:,} |".replace(",", "."))
    linhas.append(f"| Amostra Final | **{meta['amostra_final']:,}** |".replace(",", "."))
    linhas.append(f"| Nível de Confiança | {meta['confianca_pct']}% |")
    linhas.append(f"| Margem de Erro | ±{meta['margem_erro_pct']}% |")
    linhas.append(f"| Zonas Eleitorais | {meta['n_zonas']} |")
    linhas.append("")
    
    ibge = meta.get("ibge", {})
    if ibge:
        def fmt_int_opt(valor):
            if valor is None or str(valor).strip() == "":
                return "N/D"
            try:
                return f"{int(float(valor)):,}".replace(",", ".")
            except (TypeError, ValueError):
                return "N/D"

        linhas.append("## Perfil Socioeconômico (IBGE)\n")
        linhas.append("| Indicador | Valor |")
        linhas.append("|-----------|-------|")
        linhas.append(f"| População Total | {fmt_int_opt(ibge.get('POPULACAO_TOTAL'))} |")
        linhas.append(f"| ID IBGE | {fmt_int_opt(ibge.get('ID_IBGE'))} |")
        linhas.append(f"| IDH Municipal | {ibge.get('IDH') if ibge.get('IDH') not in [None, ''] else 'N/D'} |")
        if ibge.get("PIB_PER_CAPITA") not in [None, ""]:
            linhas.append(f"| PIB per Capita | R$ {float(ibge.get('PIB_PER_CAPITA')):,.2f} |".replace(",", "."))
        linhas.append("")
    
    linhas.append("## Distribuição por Zona Eleitoral\n")
    linhas.append("| Zona | Eleitores | Fem. | Masc. | Seções | Quota | Q.Fem. | Q.Masc. | % |")
    linhas.append("|------|-----------|------|-------|--------|-------|--------|---------|---|")
    
    total_el = df["ELEITORES_TOTAL"].sum()
    for _, row in df.iterrows():
        pct = f"{row['ELEITORES_TOTAL']/total_el*100:.1f}%"
        linhas.append(
            f"| {row['ZONA']} | {int(row['ELEITORES_TOTAL']):,} | {int(row['ELEITORES_FEMININO']):,} | "
            f"{int(row['ELEITORES_MASCULINO']):,} | {int(row['SECOES'])} | "
            f"**{int(row['QUOTA'])}** | {int(row['QUOTA_FEMININO'])} | {int(row['QUOTA_MASCULINO'])} | {pct} |"
            .replace(",", ".")
        )
    
    linhas.append(
        f"| **TOTAL** | **{int(df['ELEITORES_TOTAL'].sum()):,}** | {int(df['ELEITORES_FEMININO'].sum()):,} | "
        f"{int(df['ELEITORES_MASCULINO'].sum()):,} | {int(df['SECOES'].sum())} | "
        f"**{int(df['QUOTA'].sum())}** | {int(df['QUOTA_FEMININO'].sum())} | {int(df['QUOTA_MASCULINO'].sum())} | 100% |"
        .replace(",", ".")
    )
    
    linhas.append("\n---\n")
    if benchmark and benchmark.get("tabelas"):
        linhas.append("## Benchmark Estratificado\n")
        linhas.append(benchmark.get("metodologia", ""))
        linhas.append("")
        for tabela in benchmark.get("tabelas", []):
            linhas.append(f"### {tabela.get('titulo','')}\n")
            linhas.append(f"Fonte: {tabela.get('fonte','')}\n")
            linhas.append("| Categoria | V. Absoluto | % |")
            linhas.append("|-----------|-------------|---|")
            for item in tabela.get("linhas", []):
                linhas.append(
                    f"| {item.get('categoria','')} | {int(item.get('v_absoluto',0))} | {float(item.get('pct',0)):.2f}% |"
                )
            t = tabela.get("total", {})
            linhas.append(
                f"| **TOTAL** | **{int(t.get('v_absoluto',0))}** | **100,00%** |"
            )
            linhas.append("")

        if benchmark.get("observacoes"):
            linhas.append("Observações: " + " | ".join(benchmark.get("observacoes", [])))
            linhas.append("")

        linhas.append("\n---\n")

    linhas.append("## Nota Metodológica\n")
    linhas.append(meta.get("metodologia_institucional") or gerar_texto_metodologia_institucional(meta, benchmark=benchmark))
    
    with open(caminho, "w", encoding="utf-8") as f:
        f.write("\n".join(linhas))


# ─────────────────────────────────────────────────────────────────────────────
# FUNÇÕES AUXILIARES
# ─────────────────────────────────────────────────────────────────────────────

def listar_municipios(uf: str = None) -> list:
    """Lista municípios disponíveis, opcionalmente filtrado por UF."""
    caminho_ibge = os.path.join(DADOS_DIR, "ibge.csv")
    df = pd.read_csv(caminho_ibge, encoding="utf-8-sig")

    if uf:
        df = df[df["UF"] == uf.upper()]

    return df[["UF", "MUNICIPIO"]].sort_values(["UF", "MUNICIPIO"]).to_dict("records")


def listar_ufs() -> list:
    """Lista UFs disponíveis."""
    caminho_ibge = os.path.join(DADOS_DIR, "ibge.csv")
    df = pd.read_csv(caminho_ibge, encoding="utf-8-sig")
    return sorted(df["UF"].unique().tolist())


def calcular_amostra_municipio(
    uf: str,
    municipio: str,
    confianca: float = 0.95,
    margem_erro: float = 0.05,
) -> dict:
    """
    Endpoint auxiliar: calcula e retorna a amostra recomendada para um
    município sem gerar arquivos. Usado pelo frontend para pré-preencher
    o campo de amostra assim que o usuário seleciona o município.
    """
    caminho_tse = os.path.join(DADOS_DIR, "tse.csv")
    caminho_ibge = os.path.join(DADOS_DIR, "ibge.csv")

    if not os.path.exists(caminho_tse) or not os.path.exists(caminho_ibge):
        raise FileNotFoundError(
            "Dados não encontrados. Execute 'python gerar_dados.py' primeiro."
        )

    df_tse = pd.read_csv(caminho_tse, encoding="utf-8-sig")
    df_ibge = pd.read_csv(caminho_ibge, encoding="utf-8-sig")

    uf_upper = uf.upper().strip()
    municipio_strip = municipio.strip()

    mask_tse = (df_tse["UF"] == uf_upper) & (
        df_tse["MUNICIPIO"].str.lower() == municipio_strip.lower()
    )
    df_mun = df_tse[mask_tse]

    if df_mun.empty:
        raise ValueError(f"Município '{municipio}' não encontrado no estado '{uf}'.")

    mask_ibge = (df_ibge["UF"] == uf_upper) & (
        df_ibge["MUNICIPIO"].str.lower() == municipio_strip.lower()
    )
    df_ibge_mun = df_ibge[mask_ibge]
    ibge_data = df_ibge_mun.iloc[0].to_dict() if not df_ibge_mun.empty else {}

    total_eleitores = int(df_mun["ELEITORES_TOTAL"].sum())
    n_zonas = len(df_mun)
    pop_total = int(ibge_data.get("POPULACAO_TOTAL", total_eleitores))
    idh_raw = ibge_data.get("IDH")
    try:
        idh = float(idh_raw) if idh_raw is not None and str(idh_raw).strip() != "" else None
    except (TypeError, ValueError):
        idh = None

    return calcular_amostra_recomendada(
        N_eleitores=total_eleitores,
        N_populacao=pop_total,
        n_zonas=n_zonas,
        idh=idh,
        confianca=confianca,
        margem_erro=margem_erro,
    )
